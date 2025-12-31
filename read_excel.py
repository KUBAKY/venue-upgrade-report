import openpyxl
import sys

file_path = '/Users/liyaweimacbook/Desktop/programs/其他/场馆升级分析.xlsx'

try:
    workbook = openpyxl.load_workbook(file_path, data_only=True)
    print(f"Sheet names: {workbook.sheetnames}")
    
    for sheet_name in workbook.sheetnames:
        print(f"\n--- Sheet: {sheet_name} ---")
        sheet = workbook[sheet_name]
        
        # Print all rows
        for i, row in enumerate(sheet.iter_rows(values_only=True)):
            print(row)
            
except Exception as e:
    print(f"Error reading Excel file: {e}")
